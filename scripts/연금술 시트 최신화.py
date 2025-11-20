import openpyxl

if __name__ == '__main__':
    from src import auction
    from env_var import DNF_EXCEL_PATH

    excel_file = openpyxl.open(DNF_EXCEL_PATH, keep_vba=True)

    alchemy_sheet = None
    for ws in excel_file.worksheets:
        if ws.title == "연금술":
            alchemy_sheet = ws
            break
    else:
        raise Exception("연금술 시트를 찾을 수 없음")

    row_num = 3
    while True:
        try:
            item_name = ws[f"N{row_num}"].value
        except (KeyError, ValueError):
            item_name = None
        if not item_name:
            break
        previous_item_price = int(ws[f"O{row_num}"].value or "0")

        items = auction.search_by_item_name(item_name, limit=10)
        if not items:
            print(f"{item_name} 검색 결과 없음")
        else:
            print(f"{item_name} 검색 완료")
            subsum = 0
            for item in items:
                subsum += item.count
                if subsum >= 200:
                    print(f"{previous_item_price} -> {item.price}골드로 업데이트")
                    ws[f"O{row_num}"].value = item.price
                    break
        print()

        row_num += 1

    excel_file.save(DNF_EXCEL_PATH)
